#!/usr/bin/env python3
"""process_afida.py — AFIDA raw holdings xlsx -> compact county/national JSON.

Input : a directory of afida_<YEAR>.xlsx files, downloaded by hand from
        fsa.usda.gov (the server tarpits datacenter IPs — probe-verified
        2026-07-18 — so a browser download once a year IS the pipeline).
Output: data/afida/county.json    {fips: {n,st, y:{year:acres}, latest:{...}}}
        data/afida/national.json  totals by year, by country, by land type,
                                  top counties, methodology notes

File anatomy (all years 2010-2024): row1 banner, row2 spanning group labels,
row3 REAL headers, row4+ one row per holding (owner x parcel).

Methodology, validated against the official reports:
  - National total = plain sum of "Number of Acres" over all rows. The 2023
    file sums to 45,850,252 ac vs the official report's ~45.8M — FSA counts
    rows (a handful of multi-owner parcels double-count ~0.05%; 37 parcels
    of 47,006 in 2023). We match the official method and say so on-page.
  - Land types: per-row Crop/Pasture/Forest/Other Agriculture/Other Non-Ag
    acre columns (forest is a large share — the page must say "much of this
    is timberland", it's the honest headline nobody prints).
  - Country: per-row Country. "United Kingdom*"-style footnote stars stripped.
  - FIPS zero-padded to 5. Rows without a usable FIPS are kept in national
    totals but dropped from the county map (counted + reported).
Missing years (2015, 2021 not published/downloaded) stay missing — the chart
shows a gap, never a line across it. Same doctrine as cash-rent 2015.
"""
import json
import os
import re
import sys
from collections import defaultdict
from openpyxl import load_workbook

RAW = sys.argv[1] if len(sys.argv) > 1 else "/home/claude/afida/raw"
OUT = sys.argv[2] if len(sys.argv) > 2 else "data/afida"

LAND_COLS = ["Crop", "Pasture", "Forest", "Other Agriculture", "Other Non-Ag"]


def norm_country(c):
    c = str(c or "").strip().rstrip("*").strip()
    return c or "Unknown"


def read_year(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # banner
    next(rows)  # group labels
    hdr = [str(h).strip() if h else "" for h in next(rows)]
    ix = {h: i for i, h in enumerate(hdr)}
    need = ["State", "County", "FIPS", "Country", "Number of Acres"]
    for n in need:
        if n not in ix:
            raise SystemExit(f"{path}: missing column {n!r} — headers: {hdr[:12]}")
    li = {c: ix.get(c) for c in LAND_COLS}
    out = []
    for r in rows:
        if r is None or r[ix["Number of Acres"]] is None:
            continue
        try:
            ac = float(r[ix["Number of Acres"]])
        except (TypeError, ValueError):
            continue
        if ac <= 0:
            continue
        f = r[ix["FIPS"]]
        try:
            fips = f"{int(f):05d}"
        except (TypeError, ValueError):
            fips = None
        land = {}
        for c, i in li.items():
            if i is None:
                continue
            try:
                v = float(r[i] or 0)
            except (TypeError, ValueError):
                v = 0.0
            if v > 0:
                land[c] = v
        out.append({
            "fips": fips,
            "st": str(r[ix["State"]] or "").strip(),
            "co": str(r[ix["County"]] or "").strip(),
            "country": norm_country(r[ix["Country"]]),
            "ac": ac,
            "land": land,
        })
    return out


def main():
    files = sorted(f for f in os.listdir(RAW) if re.match(r"afida_\d{4}\.xlsx$", f))
    if not files:
        raise SystemExit(f"no afida_<year>.xlsx files in {RAW}")
    years = [int(re.search(r"\d{4}", f).group()) for f in files]
    print(f"processing {len(files)} years: {years}")

    county = {}          # fips -> {"n","st","y":{year:ac}}
    nat_by_year = {}     # year -> total ac
    latest = max(years)
    latest_rows = None
    no_fips = defaultdict(float)

    for f, yr in zip(files, years):
        rows = read_year(os.path.join(RAW, f))
        tot = sum(r["ac"] for r in rows)
        nat_by_year[yr] = round(tot)
        print(f"  {yr}: {len(rows):,} holdings, {tot:,.0f} ac")
        for r in rows:
            if not r["fips"]:
                no_fips[yr] += r["ac"]
                continue
            c = county.setdefault(r["fips"], {"n": r["co"], "st": r["st"], "y": {}})
            c["y"][yr] = round(c["y"].get(yr, 0) + r["ac"])
        if yr == latest:
            latest_rows = rows

    # latest-year detail per county: top countries + land split
    by_cty_country = defaultdict(lambda: defaultdict(float))
    by_cty_land = defaultdict(lambda: defaultdict(float))
    nat_country = defaultdict(float)
    nat_land = defaultdict(float)
    for r in latest_rows:
        nat_country[r["country"]] += r["ac"]
        for k, v in r["land"].items():
            nat_land[k] += v
        if r["fips"]:
            by_cty_country[r["fips"]][r["country"]] += r["ac"]
            for k, v in r["land"].items():
                by_cty_land[r["fips"]][k] += v

    for fips, c in county.items():
        cc = by_cty_country.get(fips)
        if cc:
            top = sorted(cc.items(), key=lambda kv: -kv[1])[:5]
            c["top"] = [[k, round(v)] for k, v in top]
        ll = by_cty_land.get(fips)
        if ll:
            c["land"] = {k: round(v) for k, v in sorted(ll.items(), key=lambda kv: -kv[1])}

    top_counties = sorted(
        ((f, c["n"], c["st"], c["y"].get(latest, 0)) for f, c in county.items()),
        key=lambda t: -t[3])[:25]

    os.makedirs(OUT, exist_ok=True)
    json.dump(county, open(os.path.join(OUT, "county.json"), "w"),
              separators=(",", ":"))
    national = {
        "latest_year": latest,
        "years": sorted(nat_by_year),
        "missing_years": sorted(set(range(min(years), latest + 1)) - set(years)),
        "total_by_year": nat_by_year,
        "by_country": {k: round(v) for k, v in
                       sorted(nat_country.items(), key=lambda kv: -kv[1])[:40]},
        "by_land_type": {k: round(v) for k, v in
                         sorted(nat_land.items(), key=lambda kv: -kv[1])},
        "top_counties": [{"fips": f, "n": n, "st": s, "ac": round(a)}
                         for f, n, s, a in top_counties],
        "counties_with_holdings": sum(1 for c in county.values() if c["y"].get(latest)),
        "no_fips_acres_by_year": {y: round(v) for y, v in sorted(no_fips.items())},
        "method": ("Row-sum of Number of Acres, matching FSA's official totals "
                   "(2023 check: 45,850,252 vs report ~45.8M). AFIDA is "
                   "self-reported disclosure data; leaseholds of 10+ years are "
                   "included, and a large share of the acreage is forest."),
        "source": "USDA FSA AFIDA detailed holdings files, downloaded by hand",
    }
    json.dump(national, open(os.path.join(OUT, "national.json"), "w"),
              separators=(",", ":"), indent=None)
    print(f"\nwrote {OUT}/county.json ({len(county):,} counties) and national.json")
    print(f"latest ({latest}) total: {nat_by_year[latest]:,} ac across "
          f"{national['counties_with_holdings']:,} counties")
    print("top countries:", list(national["by_country"])[:6])
    print("land types:", national["by_land_type"])
    print("missing years (honest gaps):", national["missing_years"])


if __name__ == "__main__":
    main()
