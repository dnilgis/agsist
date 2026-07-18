#!/usr/bin/env python3
"""probe_viral.py — data probes for the five viral pages. READS ONLY, writes nothing.

Run in GitHub Actions (sandbox IPs are blocked by several of these hosts) via
probe-viral.yml, then paste the log back. The fetchers and pages get written
against these results — probes before pipelines, always.

What it settles, per page:

  1. AFIDA foreign-farmland map
     - downloads the real "detailed holdings" Excel for 2024 and 2015
     - enumerates ACTUAL column names (the field-key PDF describes them, but
       the file is the truth), row counts, county identifier format,
       country-of-investor field, acres field, land-type breakdown
     - answers: can we aggregate to county FIPS? do older years share the
       schema (trend feasibility 2010->2024)?

  2. FSA payment files (the farmer-first subsidy page)
     - scrapes the FOIA payment-files page for real download links
     - downloads ONE state file for the newest year, enumerates columns
     - answers: is county derivable (county code vs address city/state)?
       is program name per-row? what does one year/state weigh?

  4. Basis vindication + 8. Bushel's journey (AgTransport Socrata)
     - re-verifies the four datasets probed 2026-07-16 are still live and
       current: grain_basis v85y-3hep, barge_rates deqi-uken,
       grain_price_spreads an4w-mnp7, transport_cost_idx 8uye-ieij
     - prints newest row date for each (freshness), row deltas vs July probe

  7. Conditions percentile
     - re-verifies weekly NASS conditions short_desc strings with the real
       key, prints newest week_ending for IA corn (freshness)

Every failure prints a verdict line instead of crashing the whole probe —
a dead source is a finding, not an error.
"""
import io
import json
import os
import re
import sys
import urllib.request
import urllib.error

UA = {"User-Agent": "AGSIST/1.0 (+https://agsist.com; data probe)"}
TIMEOUT = 120


def get(url, timeout=TIMEOUT, binary=False):
    req = urllib.request.Request(url, headers=UA)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        body = r.read()
    return body if binary else body.decode("utf-8", "replace")


def head_size(url):
    req = urllib.request.Request(url, headers=UA, method="HEAD")
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return int(r.headers.get("Content-Length") or 0), r.headers.get("Content-Type", "")
    except Exception as e:
        return None, str(e)[:60]


def sect(title):
    print("\n" + "=" * 78 + "\n  " + title + "\n" + "=" * 78)


def xlsx_survey(blob, label, max_rows=5000):
    """Column names + sample stats from an xlsx without loading everything."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            print(f"    sheet '{ws.title}': EMPTY")
            continue
        header = [str(h).strip() if h is not None else "" for h in header]
        print(f"    sheet '{ws.title}': {ws.max_row - 1 if ws.max_row else '?'} data rows (reported)")
        print(f"    columns ({len(header)}): {header}")
        # sample the first rows for the fields we care about
        sample = []
        for i, row in enumerate(rows):
            if i >= max_rows:
                break
            sample.append(row)
        if sample:
            print(f"    sampled {len(sample)} rows; first row: "
                  + json.dumps([str(v)[:28] for v in sample[0]], ensure_ascii=False))
            # per-column non-null coverage on the sample
            cover = []
            for ci, name in enumerate(header):
                nn = sum(1 for r in sample if ci < len(r) and r[ci] not in (None, ""))
                cover.append(f"{name or 'col'+str(ci)}={100*nn//len(sample)}%")
            print("    non-null coverage (sample): " + ", ".join(cover[:18])
                  + (" …" if len(cover) > 18 else ""))
        break  # first sheet is the data sheet in these files; note if more exist
    if len(wb.sheetnames) > 1:
        print(f"    NOTE: {len(wb.sheetnames)} sheets present: {wb.sheetnames}")


# ---------------------------------------------------------------- 1. AFIDA
def probe_afida():
    sect("1. AFIDA — foreign holdings detailed data (fsa.usda.gov)")
    candidates = {
        2024: "https://www.fsa.usda.gov/documents/afida-yr2024-holdings-data",
        2023: "https://www.fsa.usda.gov/documents/afida-yr2023-holdings-data",
        2015: "https://www.fsa.usda.gov/sites/default/files/documents/afida_current_holdings_yr2015.xlsx",
    }
    # also scrape the index page for every year link we can find
    try:
        idx = get("https://www.fsa.usda.gov/resources/economic-policy-analysis/afida/annual-reports-underlying-data")
        links = re.findall(r'href="([^"]*(?:afida[^"]*(?:holdings|data)[^"]*|holdings[^"]*afida[^"]*))"', idx, re.I)
        print(f"  index page: {len(links)} candidate data links found")
        for l in sorted(set(links))[:20]:
            print("    " + (l if l.startswith("http") else "https://www.fsa.usda.gov" + l))
    except Exception as e:
        print(f"  index page scrape FAILED: {type(e).__name__} {str(e)[:80]}")

    for yr, url in candidates.items():
        try:
            size, ctype = head_size(url)
            print(f"\n  YR{yr}: HEAD {size if size else '?'}B  type={ctype}")
            blob = get(url, binary=True)
            print(f"  YR{yr}: downloaded {len(blob):,}B — surveying")
            xlsx_survey(blob, f"afida-{yr}")
        except Exception as e:
            print(f"  YR{yr}: FAILED {type(e).__name__}: {str(e)[:100]}")


# ---------------------------------------------------------- 6. FSA payments
def probe_fsa_payments():
    sect("6. FSA payment files (FOIA reading room)")
    page_urls = [
        "https://www.fsa.usda.gov/tools/informational/freedom-information-act-foia/electronic-reading-room/frequently-requested/payment-files",
    ]
    links = []
    for pu in page_urls:
        try:
            html = get(pu)
            found = re.findall(r'href="([^"]+)"[^>]*>([^<]{4,80})</a>', html)
            data_links = [(u, t.strip()) for u, t in found
                          if re.search(r"payment|name.?address", t, re.I)
                          and re.search(r"documents/|\.xlsx|\.zip", u, re.I)]
            print(f"  {pu.split('/')[-1]}: {len(data_links)} payment-file links")
            for u, t in data_links[:25]:
                full = u if u.startswith("http") else "https://www.fsa.usda.gov" + u
                links.append((full, t))
                print(f"    {t[:60]:<60} {full[:80]}")
        except Exception as e:
            print(f"  page scrape FAILED: {type(e).__name__} {str(e)[:80]}")

    # download the SMALLEST discovered file to survey the schema politely
    best = None
    for full, t in links:
        size, _ = head_size(full)
        if size:
            if best is None or size < best[0]:
                best = (size, full, t)
    if best:
        size, full, t = best
        print(f"\n  downloading smallest file for schema survey: {t} ({size:,}B)")
        try:
            blob = get(full, binary=True)
            xlsx_survey(blob, "fsa-payments")
        except Exception as e:
            print(f"  download FAILED: {type(e).__name__} {str(e)[:100]}")
    else:
        print("  no downloadable file found — schema survey skipped")


# ------------------------------------------- 4+8. AgTransport re-verification
def probe_agtransport():
    sect("4+8. AgTransport (Socrata, no key) — freshness re-check vs 2026-07-16 probe")
    JULY_PROBE = {"v85y-3hep": 27582, "g92w-8cn7": 38410,
                  "an4w-mnp7": 34658, "deqi-uken": 8225, "8uye-ieij": 1246}
    names = {"v85y-3hep": "grain_basis", "g92w-8cn7": "grain_prices",
             "an4w-mnp7": "grain_price_spreads", "deqi-uken": "barge_rates",
             "8uye-ieij": "transport_cost_idx"}
    for ds, prev in JULY_PROBE.items():
        try:
            cnt = json.loads(get(f"https://agtransport.usda.gov/resource/{ds}.json?$select=count(*)"))[0]
            n = int(list(cnt.values())[0])
            newest = json.loads(get(
                f"https://agtransport.usda.gov/resource/{ds}.json?$select=max(date)"))[0]
            nd = list(newest.values())[0]
            print(f"  {names[ds]:<20} {ds}  rows={n:,} (Δ{n-prev:+,} since 7/16)  newest={str(nd)[:10]}")
        except Exception as e:
            print(f"  {names[ds]:<20} {ds}  FAILED {type(e).__name__}: {str(e)[:80]}")


# --------------------------------------------------- 7. Conditions freshness
def probe_conditions():
    sect("7. NASS weekly conditions — freshness re-check")
    key = os.environ.get("NASS_API_KEY", "").strip()
    if not key:
        print("  NASS_API_KEY not set — skipped (set the secret in probe-viral.yml)")
        return
    import urllib.parse
    q = urllib.parse.urlencode({
        "key": key,
        "short_desc": "CORN - CONDITION, MEASURED IN PCT EXCELLENT",
        "agg_level_desc": "STATE", "state_alpha": "IA",
        "year": "2026", "format": "JSON",
    })
    try:
        d = json.loads(get("https://quickstats.nass.usda.gov/api/api_GET/?" + q))
        rows = d.get("data", [])
        weeks = sorted(r.get("week_ending", "") for r in rows)
        print(f"  IA corn PCT EXCELLENT 2026: {len(rows)} weekly rows, "
              f"first={weeks[0] if weeks else '?'} newest={weeks[-1] if weeks else '?'}")
        if rows:
            print(f"  newest value: {rows[-1].get('Value')}% excellent "
                  f"(week ending {rows[-1].get('week_ending')})")
    except Exception as e:
        print(f"  FAILED {type(e).__name__}: {str(e)[:100]}")


def main():
    print("PROBE VIRAL — reads only, writes nothing")
    print("Settles data questions for: AFIDA map, FSA payments page, basis page,")
    print("conditions percentile, bushel's journey. Paste this log back.")
    probe_afida()
    probe_fsa_payments()
    probe_agtransport()
    probe_conditions()
    print("\nDONE — paste the full log back and the fetchers get written against it.")


if __name__ == "__main__":
    main()
